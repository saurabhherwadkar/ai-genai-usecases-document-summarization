# Test excel loader module - unit tests for the ExcelLoader class.
# Tests Excel file loading, worksheet processing, and error handling.

import pytest
from openpyxl import Workbook

from src.ingestion.excel_loader import ExcelLoader
from src.utils.exceptions import IngestionError


class TestExcelLoader:
    """Tests for the ExcelLoader class."""

    def setup_method(self):
        """Set up test fixtures for each test method."""
        # Create a fresh loader instance for each test
        self.loader = ExcelLoader()

    def test_load_valid_xlsx_file(self, tmp_path):
        """Test that a valid XLSX file is loaded and converted to text."""
        # Create a test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(["Project Name", "Location", "Budget"])
        ws.append(["Office Tower", "Chicago, IL", "$45M"])
        ws.append(["Bridge Repair", "Springfield, IL", "$12M"])
        excel_path = tmp_path / "test_projects.xlsx"
        wb.save(str(excel_path))

        # Load the file
        content = self.loader.load(str(excel_path))

        # Verify content contains expected data
        assert "Projects" in content
        assert "Office Tower" in content
        assert "Chicago, IL" in content
        assert "$45M" in content
        assert "Bridge Repair" in content

    def test_load_multiple_worksheets(self, tmp_path):
        """Test that all worksheets are included in the output."""
        # Create a workbook with multiple sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Permits"
        ws1.append(["Permit ID", "Status"])
        ws1.append(["BP-001", "Approved"])

        ws2 = wb.create_sheet("Contractors")
        ws2.append(["Name", "Phone"])
        ws2.append(["ABC Corp", "555-0100"])
        excel_path = tmp_path / "multi_sheet.xlsx"
        wb.save(str(excel_path))

        # Load the file
        content = self.loader.load(str(excel_path))

        # Verify both sheets are represented
        assert "Permits" in content
        assert "Contractors" in content
        assert "BP-001" in content
        assert "ABC Corp" in content

    def test_load_nonexistent_file_raises_error(self):
        """Test that loading a nonexistent file raises IngestionError."""
        with pytest.raises(IngestionError) as exc_info:
            self.loader.load("/nonexistent/file.xlsx")

        assert "not found" in exc_info.value.message.lower()

    def test_load_unsupported_extension_raises_error(self, tmp_path):
        """Test that a non-Excel extension raises IngestionError."""
        # Create a file with wrong extension
        fake_file = tmp_path / "data.csv"
        fake_file.write_text("col1,col2", encoding="utf-8")

        with pytest.raises(IngestionError) as exc_info:
            self.loader.load(str(fake_file))

        assert "Unsupported Excel format" in exc_info.value.message

    def test_load_empty_workbook_raises_error(self, tmp_path):
        """Test that an empty workbook raises IngestionError."""
        # Create an empty workbook (no data rows)
        wb = Workbook()
        ws = wb.active
        # Don't add any rows
        excel_path = tmp_path / "empty.xlsx"
        wb.save(str(excel_path))

        with pytest.raises(IngestionError) as exc_info:
            self.loader.load(str(excel_path))

        assert "No data found" in exc_info.value.message

    def test_load_handles_none_cell_values(self, tmp_path):
        """Test that None cell values are handled gracefully."""
        # Create a workbook with some empty cells
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Budget", "Phase"])
        ws.append(["Project A", None, "Planning"])
        ws.append([None, "$5M", None])
        excel_path = tmp_path / "sparse.xlsx"
        wb.save(str(excel_path))

        # Load should succeed without errors
        content = self.loader.load(str(excel_path))

        # Verify content was extracted
        assert "Project A" in content
        assert "$5M" in content
        # N/A should appear for None cells
        assert "N/A" in content
