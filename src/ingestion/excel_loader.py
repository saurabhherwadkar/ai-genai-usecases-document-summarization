# Excel loader module - loads and converts Excel spreadsheets to text format.
# Handles multiple worksheets, tables, and structured data for RAG ingestion.

from pathlib import Path

from openpyxl import load_workbook

from src.utils.exceptions import IngestionError
from src.utils.logger import get_logger

# Module logger for tracking Excel loading events
logger = get_logger(__name__)


class ExcelLoader:
    """Loads and converts Excel files to text representation for RAG ingestion.

    Iterates through all worksheets in an Excel file, converting each row
    into a readable text format suitable for chunking and embedding.
    """

    def load(self, file_path: str) -> str:
        """Load an Excel file and convert all worksheets to text.

        Each worksheet is converted to a text table with headers and rows.
        Multiple worksheets are separated by section headers.

        Args:
            file_path: Absolute or relative path to the Excel file.

        Returns:
            str: Text representation of all worksheet data.

        Raises:
            IngestionError: If the file does not exist, cannot be parsed, or is empty.
        """
        # Convert the path string to a Path object for validation
        path = Path(file_path)

        # Verify that the file exists on disk
        if not path.exists():
            raise IngestionError(f"Excel file not found: {file_path}", details={"path": file_path})

        # Verify the file extension is a supported Excel format
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise IngestionError(
                f"Unsupported Excel format: {path.suffix}",
                details={"path": file_path, "extension": path.suffix},
            )

        # Log the loading attempt
        logger.info("Loading Excel file: %s", path.name)

        try:
            # Load the workbook in read-only mode for memory efficiency
            workbook = load_workbook(str(path), read_only=True, data_only=True)

            # Process each worksheet and collect text output
            all_sheets_text = []
            for sheet_name in workbook.sheetnames:
                # Convert each worksheet to text format
                sheet_text = self._process_worksheet(workbook[sheet_name], sheet_name)
                if sheet_text:
                    all_sheets_text.append(sheet_text)

            # Close the workbook to release file handle
            workbook.close()

            # Verify that some data was extracted
            if not all_sheets_text:
                raise IngestionError(f"No data found in Excel file: {path.name}", details={"path": file_path})

            # Join all worksheet texts with double newlines
            combined_text = "\n\n".join(all_sheets_text)

            # Log successful loading with output length
            logger.info("Successfully extracted %d characters from Excel file %s", len(combined_text), path.name)

            return combined_text

        except IngestionError:
            # Re-raise IngestionErrors without wrapping
            raise
        except Exception as error:
            # Wrap unexpected errors in IngestionError
            raise IngestionError(
                f"Failed to load Excel file: {path.name}",
                details={"path": file_path, "error": str(error)},
            ) from error

    def _process_worksheet(self, worksheet, sheet_name: str) -> str:
        """Convert a single worksheet to a text representation.

        Reads all rows from the worksheet and formats them as a text table
        with headers derived from the first row.

        Args:
            worksheet: The openpyxl worksheet object to process.
            sheet_name: Name of the worksheet for the section header.

        Returns:
            str: Text representation of the worksheet data, or empty string if no data.
        """
        # Collect all rows from the worksheet
        rows = list(worksheet.iter_rows(values_only=True))

        # Return empty string if worksheet has no rows
        if not rows:
            logger.debug("Worksheet '%s' is empty, skipping", sheet_name)
            return ""

        # Filter out completely empty rows (all None values)
        non_empty_rows = [row for row in rows if any(cell is not None for cell in row)]

        # Return empty string if all rows are empty
        if not non_empty_rows:
            logger.debug("Worksheet '%s' has no data rows, skipping", sheet_name)
            return ""

        # Use the first row as column headers
        headers = [str(cell) if cell is not None else "" for cell in non_empty_rows[0]]

        # Build the text output starting with a section header
        text_parts = [f"=== Sheet: {sheet_name} ==="]

        # Add headers line
        text_parts.append("Columns: " + " | ".join(headers))
        text_parts.append("-" * 40)

        # Process data rows (skip header row)
        for row_index, row in enumerate(non_empty_rows[1:], start=1):
            # Format each cell value, pairing with its header
            row_parts = []
            for header, cell in zip(headers, row):
                # Convert cell value to string, defaulting empty cells to "N/A"
                cell_value = str(cell) if cell is not None else "N/A"
                row_parts.append(f"{header}: {cell_value}")
            # Join cell pairs with separator and add to output
            text_parts.append(f"Row {row_index}: " + " | ".join(row_parts))

        # Log the number of data rows processed
        logger.debug("Processed %d data rows from worksheet '%s'", len(non_empty_rows) - 1, sheet_name)

        # Join all text parts with newlines
        return "\n".join(text_parts)
